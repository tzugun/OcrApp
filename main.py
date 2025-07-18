# main.py

import os
from datetime import datetime
from kivy.app import App
from kivy.clock import Clock
from kivy.lang import Builder
from kivy.properties import ListProperty, StringProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.core.audio import SoundLoader

import cv2, numpy as np
from pyzbar.pyzbar import decode
import pytesseract
from openpyxl import Workbook, load_workbook

KV = '''
<RootWidget>:
    orientation: 'vertical'

    Camera:
        id: cam
        resolution: (640, 480)
        play: True

    BoxLayout:
        size_hint_y: None
        height: dp(30)
        Label:
            text: "Last Scan: " + root.last_scan

    Spinner:
        id: history_spinner
        text: "Scan History"
        values: root.history_items
        size_hint_y: None
        height: dp(44)
'''

class RootWidget(BoxLayout):
    last_scan = StringProperty('—')
    history_items = ListProperty([])

class QRLoggerApp(App):
    def build(self):
        Builder.load_string(KV)
        self.root = RootWidget()
        self.sound = SoundLoader.load('assets/beep.wav')
        self._prepare_excel()
        Clock.schedule_interval(self._scan_loop, 1/10.)
        return self.root

    def _prepare_excel(self):
        docs = os.path.join(self.user_data_dir, '..', 'Documents')
        os.makedirs(docs, exist_ok=True)
        self.xlsx = os.path.join(docs, 'qr_scans.xlsx')
        if not os.path.exists(self.xlsx):
            wb = Workbook()
            ws = wb.active
            ws.append(['Timestamp','Patient Name','Order Number','Product Name','Power','Quantity'])
            wb.save(self.xlsx)

    def _scan_loop(self, dt):
        cam = self.root.ids.cam
        if not cam.texture:
            return

        buf = np.frombuffer(cam.texture.pixels, np.uint8)
        img = cv2.imdecode(buf, cv2.IMREAD_COLOR)
        codes = decode(img)
        if not codes:
            return

        code = codes[0]
        order = code.data.decode()
        # Crop around QR for patient-name OCR
        pts = code.polygon
        xs, ys = [p.x for p in pts], [p.y for p in pts]
        x1 = max(min(xs) - 30, 0)
        x2 = min(max(xs) + 30, img.shape[1])
        y1 = max(min(ys) - 30, 0)
        y2 = min(max(ys) + 30, img.shape[0])
        crop = img[y1:y2, x1:x2]
        text = pytesseract.image_to_string(crop).splitlines()
        patient = next((l for l in text if l.strip()), '')

        # Full-page OCR for table fields
        full = pytesseract.image_to_string(img)
        product, power, qty = self._parse(full)

        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self._log_excel([ts, patient, order, product, power, qty])

        # Update UI
        self.root.last_scan = f"{order} @ {ts}"
        entry = f"{order} – {patient}"
        if entry not in self.root.history_items:
            self.root.history_items.append(entry)

        # Beep on success
        if self.sound:
            self.sound.play()

    def _parse(self, text):
        product = power = qty = ''
        for line in text.splitlines():
            if 'Lens Description' in line or line.strip().startswith('-'):
                product = line.strip()
            if 'Power' in line:
                for w in line.split():
                    if w.startswith('-') or w.replace('.','',1).isdigit():
                        power = w
            if 'Qty' in line or 'Quantity' in line:
                for w in line.split():
                    if w.isdigit():
                        qty = w
        return product, power, qty

    def _log_excel(self, row):
        wb = load_workbook(self.xlsx)
        wb.active.append(row)
        wb.save(self.xlsx)

if __name__ == '__main__':
    QRLoggerApp().run()
